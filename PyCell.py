#
# This progam is using openpyxl, termcolor and
#

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from termcolor import colored

wb = Workbook

#Strona główna
def MainPage():
    main = colored("PyCell\n", 'green')
    print(main)
    main2 = colored("Welcome to PyCell, the Python-Powered Excel Sheets Editor.\n", 'green')
    print(main2)
    main3 = colored("This program is used to edit Excel sheets (.xlsx). It includes many sheet settings so you can work easily and efficiently without too much editing. You can also create shortcuts for faster sheet editing.\n", 'yellow')
    print(main3)
    main4 = colored("Helpful commands:", 'yellow')
    print(main4)
    print("gen_new - Creates a new Excel .xlsx sheet file (saves to the directory where PyCell is located)\ntextcell - Runs cell editing setup\nimagecell - Runs configuration of image cells\nextended-cell - Launches advanced cell editing configuration \ndelete - Deletes the given sheet \nmigrate - Runs configuration of settings between different sheets\nautomate - Runs configuration of Automate Shortcuts* function\nsave - Saves all data (When you close program, you will lose all work progress)\nread-saved - Read and load all saved data\n")
    print("For more commands, please type help()\n")
    main5 = ("Meaning for Automate Shortcuts:", 'yellow')
    print("Automate Shortcuts - It allows you to organize tasks faster. Use the automate command to start the creative process. The function works in such a way that it creates commands that will be displayed on the command screen (when starting the program) and allows for automatic execution of the task that was set during configuration. To start editing sheets, create a new file using the gen-new command. The configuration will guide you through a series of several settings to get everything working properly.\n")
    userAction = input(colored("[PyCell] > ", 'green'))
    print(userAction)

MainPage()

userAction = input(colored("[PyCell] > ", 'green'))

endsuccess = colored("[PyCell Setup] Operation ended successfully", 'green')

def gen_new():
    newConfig = input(colored("\n[PyCell Setup] Okay, tell me name of your .xlsx file:\n[PyCell Setup] > ", 'blue'))
    print(newConfig)
    wb = openpyxl.Workbook()
    sheet = wb.active
    wb.create_sheet(newConfig)
    wb.save(filename = newConfig + ".xlsx")
    print(endsuccess)
    MainPage()

def textcell():
    wb = load_workbook(filename = newConfig + ".xlsx")
    newtc3 = input(colored("\n[PyCell Setup] Choose number of row to continue:\n[PyCell Setup] > ", 'blue'))
    print(newtc3)
    newtc4 = input(colored("\n[PyCell Setup] Choose number of column to complete the first step:\n[PyCell Setup] > ", 'blue'))
    print(newtc4)
    newtc5 = input(colored("[PyCell Setup] Okay, now input text that will appear:\n[PyCell Setup] > ", 'blue'))
    print(newtc5)
    c1 = sheet.cell(row = newtc, column = newtc2)
    c1.value = newtc5
    print(endsuccess)
    MainPage()


if userAction == "gen_new":
    gen_new()

if userAction == "textcell":
    textcell()


